# From Python
# It requires OpenCV installed for Python
import sys
import cv2
import os
from sys import platform
import argparse
import numpy as np #new added
import openpyxl
from openpyxl import Workbook

def get_latest_gallery_directory(base_path):
    # Find the latest gallery directory number
    i = 1
    latest_gallery = None
    while True:
        gallery_path = os.path.join(base_path, f'frameGallery{i}')
        if not os.path.exists(gallery_path):
            break
        latest_gallery = gallery_path
        i += 1
    return latest_gallery

def get_new_filename(base_path):
    i = 1
    while True:
        filename = os.path.join(base_path, f"frame{i}.jpg")
        if not os.path.exists(filename):
            return filename
        i += 1
       
def write_keypoints_to_excel(right_buffer_coords, left_buffer_coords, saved_filename):
    base_directory = "../../../yolov7/inference/detectImg"
    latest_gallery_directory = get_latest_gallery_directory(base_directory)

     # Construct the path to the latest Excel file
    excel_file_path = os.path.join(latest_gallery_directory, 'keypoint_coordinates.xlsx')

    # Load the existing Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Find the last row in the sheet
    last_row = sheet.max_row + 1

    # Write the buffer coordinates to the Excel sheet
    sheet.cell(row=last_row, column=1, value=saved_filename)
    sheet.cell(row=last_row, column=2, value=right_buffer_coords[0])
    sheet.cell(row=last_row, column=3, value=right_buffer_coords[1])
    sheet.cell(row=last_row, column=4, value=right_buffer_coords[2])
    sheet.cell(row=last_row, column=5, value=right_buffer_coords[3])
    sheet.cell(row=last_row, column=6, value=left_buffer_coords[0])
    sheet.cell(row=last_row, column=7, value=left_buffer_coords[1])
    sheet.cell(row=last_row, column=8, value=left_buffer_coords[2])
    sheet.cell(row=last_row, column=9, value=left_buffer_coords[3])

    # Save the Excel file
    workbook.save(excel_file_path)


try:
    # Import Openpose (Windows/Ubuntu/OSX)
    dir_path = os.path.dirname(os.path.realpath(__file__))
    try:
        # Windows Import
        if platform == "win32":
            # Change these variables to point to the correct folder (Release/x64 etc.)
            sys.path.append(dir_path + '/../../build/python/openpose/Release');
            os.environ['PATH']  = os.environ['PATH'] + ';' + dir_path + '/../../build/x64/Release;' +  dir_path + '/../../build/bin;'
            import pyopenpose as op
        else:
            # Change these variables to point to the correct folder (Release/x64 etc.)
            sys.path.append('../../python');
            # If you run `make install` (default path is `/usr/local/python` for Ubuntu), you can also access the OpenPose/python module from there. This will install OpenPose and the python library at your desired installation path. Ensure that this is in your python path in order to use it.
            # sys.path.append('/usr/local/python')
            from openpose import pyopenpose as op
    except ImportError as e:
        print('Error: OpenPose library could not be found. Did you enable `BUILD_PYTHON` in CMake and have this Python script in the right folder?')
        raise e

    # Flags
    parser = argparse.ArgumentParser()
    parser.add_argument("--image_path", default="../media/COCO_val2014_000000000192.jpg", help="Process an image. Read all standard formats (jpg, png, bmp, etc.).")
    args = parser.parse_known_args()

    # Custom Params (refer to include/openpose/flags.hpp for more parameters)
    params = dict()
    params["model_folder"] = "../../models/"

    # Add others in path?
    for i in range(0, len(args[1])):
        curr_item = args[1][i]
        if i != len(args[1])-1: next_item = args[1][i+1]
        else: next_item = "1"
        if "--" in curr_item and "--" in next_item:
            key = curr_item.replace('-','')
            if key not in params:  params[key] = "1"
        elif "--" in curr_item and "--" not in next_item:
            key = curr_item.replace('-','')
            if key not in params: params[key] = next_item

    # Construct it from system arguments
    # op.init_argv(args[1])
    # oppython = op.OpenposePython()

    # Starting OpenPose
    opWrapper = op.WrapperPython()
    opWrapper.configure(params)
    opWrapper.start()

    # Process Image
    datum = op.Datum()
    imageToProcess = cv2.imread(args[0].image_path)
    datum.cvInputData = imageToProcess
    opWrapper.emplaceAndPop(op.VectorDatum([datum]))

    # Display Image
    print("Body keypoints: \n" + str(datum.poseKeypoints))
    #cv2.imshow("OpenPose 1.7.0 - Tutorial Python API", datum.cvOutputData)
    #cv2.waitKey(0)

    # Test
    # Index of the keypoint you want to extract (e.g., 0 for the first keypoint)
    right_keypoint_index = 4
    left_keypoint_index = 7

    # Extract the specified keypoint from the keypoints
    right_keypoint = datum.poseKeypoints[0][right_keypoint_index]
    left_keypoint = datum.poseKeypoints[0][left_keypoint_index]

    # Convert keypoints to integer coordinates
    right_x, right_y = int(right_keypoint[0]), int(right_keypoint[1])
    left_x, left_y = int(left_keypoint[0]), int(left_keypoint[1])

    # Print the extracted values
    print("Right Keypoint coordinate: (", right_keypoint[0], ",", right_keypoint[1], ")")
    print("Left Keypoint coordinate: (", left_keypoint[0], ",", left_keypoint[1], ")")

   # Set the buffer rectangle size and line thickness
    buffer_width = 20  # Width of the buffer rectangle
    buffer_height = 60  # Height of the buffer rectangle
    buffer_thickness = 2  # Thickness of the buffer rectangle line

    # Calculate buffer rectangle coordinates
    right_buffer_x1 = right_x - buffer_width // 2
    right_buffer_y1 = right_y - buffer_height // 2
    right_buffer_x2 = right_x + buffer_width // 2
    right_buffer_y2 = right_y + buffer_height // 2

    # Calculate buffer rectangle coordinates
    left_buffer_x1 = left_x - buffer_width // 2
    left_buffer_y1 = left_y - buffer_height // 2
    left_buffer_x2 = left_x + buffer_width // 2
    left_buffer_y2 = left_y + buffer_height // 2

    # Draw buffer rectangle on the image
    new = cv2.rectangle(datum.cvOutputData, (right_buffer_x1, right_buffer_y1), (right_buffer_x2, right_buffer_y2), (0, 0, 255), buffer_thickness)
    new_1 = cv2.rectangle(new, (left_buffer_x1, left_buffer_y1), (left_buffer_x2, left_buffer_y2), (0, 0, 255), buffer_thickness)
    #new_2 = cv2.polylines(img=new_1, pts=[np.array([[463,-1],[534,223],[138,369],[135,332]], dtype=np.int32)], isClosed=True, color=(0,255,0), thickness=3)
    #new_3 = cv2.polylines(img=new_2, pts=[np.array([[139,377],[137,423],[539,549],[534,237]], dtype=np.int32)], isClosed=True, color=(255,0,0), thickness=3)

    #cv2.imshow("Image with Buffer Rectangle", new)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

    base_directory = '../../../yolov7/inference/detectImg'
    latest_gallery_directory = get_latest_gallery_directory(base_directory)

    folder_dir = latest_gallery_directory

    # Find the latest image number in the directory
    image_number = 1
    for image_file in os.listdir(folder_dir):
        if image_file.startswith("frame") and image_file.endswith(".jpg"):
            image_number += 1

    # Set the filename and path for saving the image
    saved_filename = f"frame{image_number}.jpg"
    saved_path = os.path.join(folder_dir, saved_filename)

    # Your existing code for saving the image
    result = cv2.imwrite(saved_path, new)

    # Write keypoints to Excel file
    right_buffer_coords = (right_buffer_x1, right_buffer_y1, right_buffer_x2, right_buffer_y2)
    left_buffer_coords = (left_buffer_x1, left_buffer_y1, left_buffer_x2, left_buffer_y2)
    write_keypoints_to_excel(right_buffer_coords, left_buffer_coords, saved_filename)
    #write_keypoints_to_excel(right_keypoint, left_keypoint, saved_filename)

    if result==True:
        print("File saved successfully with file name:", saved_filename)
    else:
        print("Error in saving file")

except Exception as e:
    print(e)
    sys.exit(-1)




# C:\Users\60115\.conda\envs\ailib\aiLibrary\openpose\examples\tutorial_api_python\openposeTest.py
# C:\Users\60115\.conda\envs\ailib\aiLibrary\yolov7\detect.py
