#call and run two separate Python files from a single driver file
#from yolov7.detect import detect
#from openpose.examples.tutorial_api_python.openposeTest import openposeTest

import subprocess
import os
from os import listdir
import openpyxl


def get_latest_gallery_directory(base_path):
    # Find the latest gallery directory number
    i = 1
    latest_gallery = None
    while True:
        gallery_path = os.path.join(base_path, f'frameGallery{i}')
        if not os.path.exists(gallery_path):
            break
        latest_gallery = gallery_path
        i += 1
    return latest_gallery

def get_next_gallery_directory(base_path):
    # Find the last gallery directory number and increment it
    i = 1
    while True:
        gallery_path = os.path.join(base_path, f'frameGallery{i}')
        if not os.path.exists(gallery_path):
            return gallery_path
        i += 1

def create_empty_excel_file(base_directory):

    # Create the new Excel file
    excel_file_name = os.path.join(base_directory, f'keypoint_coordinates.xlsx')
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the header
    sheet.cell(row=1, column=1, value="Filename")
    sheet.cell(row=1, column=2, value="Right_buffer_x1")
    sheet.cell(row=1, column=3, value="Right_buffer_y1")
    sheet.cell(row=1, column=4, value="Right_buffer_x2")
    sheet.cell(row=1, column=5, value="Right_buffer_y2")
    sheet.cell(row=1, column=6, value="Left_buffer_x1")
    sheet.cell(row=1, column=7, value="Left_buffer_y1")
    sheet.cell(row=1, column=8, value="Left_buffer_x2")
    sheet.cell(row=1, column=9, value="Left_buffer_y2")

    # Save the Excel file
    workbook.save(excel_file_name)

    return excel_file_name

def yolov7():
    os.chdir("../../../yolov7")
    # creating new yolo output gallery directory
    try:
        yolo_base_directory = 'inference/detectImg'
        yolo_gallery_directory = get_latest_gallery_directory(yolo_base_directory)
        # os.makedirs(yolo_gallery_directory)
        result_directory_name = "result"
        result_directory_path = os.path.join(yolo_gallery_directory, result_directory_name)
        os.makedirs(result_directory_path)



    except OSError:
        print("Error: Creating Directory of Data")

    print("Let us start YOLOv7 detection\n")
    # Checking the lastest gallery directory which going to go through yolo detection
    yolo_base_directory = "inference/detectImg"
    yolo_latest_gallery_directory = get_latest_gallery_directory(yolo_base_directory)

    if not yolo_latest_gallery_directory:
        print("No gallery directory found.")
        return

    folder_dir = yolo_latest_gallery_directory

    for image in os.listdir(folder_dir):
        # Check if the image ends with .jpg
        if image.endswith(".jpg"):
            image_path = os.path.join(folder_dir, image)
            command1 = ["python", "detect.py", "--source", image_path]

            try:
                subprocess.run(command1, check=True)  # check=True will raise an exception if the command returns a non-zero status
            except subprocess.CalledProcessError as e:
                print(f"Error executing command for {image_path}: {e}")
    

def openpose():
    # creating new openpose output gallery directory
    try:
        openpose_base_directory = 'yolov7/inference/detectImg'
        openpose_gallery_directory = get_next_gallery_directory(openpose_base_directory)
        os.makedirs(openpose_gallery_directory)

        excel_file_name = create_empty_excel_file(openpose_gallery_directory)
        print(f"Excel file created: {excel_file_name}")
        
    except OSError:
        print("Error: Creating Directory of Data")
        
    # Change the working directory
    os.chdir("openpose/examples/tutorial_api_python")
    # Checking the lastest gallery directory which going to go through openpose detection
    openpose_detection_base_directory = '../../../yolov7/inference/video'
    openpose_latest_gallery_directory = get_latest_gallery_directory(openpose_detection_base_directory)

    if not openpose_latest_gallery_directory:
        print("No gallery directory found.")
        return

    folder_dir = openpose_latest_gallery_directory

    # iterate over files in that directory
    for image in os.listdir(folder_dir):
        # Check if the image ends with .jpg
        if image.endswith(".jpg"):
            image_path = os.path.join(folder_dir, image)
            command2 = ["python", "openposeTest.py", "--image_path", image_path]

            try:
                subprocess.run(command2, check=True)  # check=True will raise an exception if the command returns a non-zero status
            except subprocess.CalledProcessError as e:
                print(f"Error executing command for {image_path}: {e}")

openpose()
yolov7()







        













